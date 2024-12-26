import pandas as pd
import os
from datetime import datetime
import re
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import numpy as np
from matplotlib import font_manager

# 日本語フォント（MSゴシック）の設定
font_path = '/Users/nakamurakouga/Library/Fonts/msgothic.ttc'  # フォントのパスを正確に指定
font_prop = font_manager.FontProperties(fname=font_path)
plt.rcParams['font.family'] = font_prop.get_name()

# ファイルパスと列名の定義
file_paths_excel = '/Users/nakamurakouga/Library/CloudStorage/GoogleDrive-21t0079w@student.gs.chiba-u.jp/.shortcut-targets-by-id/1NV_iyHzURznfz7V3ld50oa465FHBCJRv/林立也研究室共有フォルダ/01　研究関連/ザイマックス共同研究関連/ザイマックス共同研究_研究資料_2024/プログラム開発用/filemaker.xlsx'
property_cd_df = pd.read_excel('/Users/nakamurakouga/Documents/Chiba-U/2024研究関連/1ザイマックス/分析用/客観データ/客観その他/物件名物件CD.xlsx')

column_names = [
    "1.1躯体の耐震性能", "1.2設備の信頼性", "1.3災害時エネルギー供給", "1.4自然災害リスク対策", 
    "1.5BCPの有無", "2.1セキュリティ設備", "2.2バリアフリー法への対応", "2.3土壌環境品質・ブラウンフィールド再生", 
    "1.1外観デザイン", "2.1オフィスからの眺望", "2.2生物多様性の向上", "2.3トイレの充足性・機能性", 
    "2.4リフレッシュスペース", "3.1建築物衛生基準への適合状況", "3.2自然換気性能", "3.3自然光の導入", 
    "3.4分煙対応、禁煙対応", "4.1維持管理", "4.2満足度調査の定期的実施等", "4.3健康維持・増進プログラム", 
    "1.1空間の形状・自由さ", "1.2動線における出会いの場の創出", "1.3打ち合わせスペース", 
    "2.1高度情報通信インフラ", "2.2情報共有インフラ"
]

# カテゴリとそれに対応する列名の定義
categories = {
    "防災対策": ["1.1躯体の耐震性能", "1.2設備の信頼性", "1.3災害時エネルギー供給", "1.4自然災害リスク対策", "1.5BCPの有無"],
    "安心安全対策": ["2.1セキュリティ設備", "2.2バリアフリー法への対応", "2.3土壌環境品質・ブラウンフィールド再生"],
    "デザイン性": ["1.1外観デザイン"],
    "リフレッシュ": ["2.1オフィスからの眺望", "2.2生物多様性の向上", "2.3トイレの充足性・機能性", "2.4リフレッシュスペース"],
    "室内環境質": ["3.1建築物衛生基準への適合状況", "3.2自然換気性能", "3.3自然光の導入", "3.4分煙対応、禁煙対応"],
    "維持管理・運営": ["4.1維持管理", "4.2満足度調査の定期的実施等", "4.3健康維持・増進プログラム"],
    "空間・内装": ["1.1空間の形状・自由さ", "1.2動線における出会いの場の創出", "1.3打ち合わせスペース"],
    "情報通信": ["2.1高度情報通信インフラ", "2.2情報共有インフラ"]
}

# Qw列の定義
qw_columns = {
    "Qw1安心・安全": ["1.1躯体の耐震性能", "1.2設備の信頼性", "1.3災害時エネルギー供給", "1.4自然災害リスク対策", "1.5BCPの有無", 
                     "2.1セキュリティ設備", "2.2バリアフリー法への対応", "2.3土壌環境品質・ブラウンフィールド再生"],
    "Qw2健康性・快適性": ["1.1外観デザイン", "2.1オフィスからの眺望", "2.2生物多様性の向上", "2.3トイレの充足性・機能性", 
                     "2.4リフレッシュスペース", "3.1建築物衛生基準への適合状況", "3.2自然換気性能", "3.3自然光の導入", 
                     "3.4分煙対応、禁煙対応", "4.1維持管理", "4.2満足度調査の定期的実施等", "4.3健康維持・増進プログラム"],
    "Qw3知的生産性向上": ["1.1空間の形状・自由さ", "1.2動線における出会いの場の創出", "1.3打ち合わせスペース", 
                       "2.1高度情報通信インフラ", "2.2情報共有インフラ"]
}

# ファイルパスを読み込む関数
def read_file_paths(file_path):
    df = pd.read_excel(file_path)
    return df['保存ファイルパス'].tolist()

# 物件名をファイル名から抽出する関数（半角カッコ）
def extract_property_name(file_name):
    match = re.search(r'\((.*?)\)', file_name)
    return match.group(1) if match else '不明'

# Excelファイルを処理してデータを抽出する関数
def process_excel_file_if_exists(file_path, sheet_name='２．評価結果集計シート'):
    try:
        # シートから必要な列を読み込み
        data = pd.read_excel(file_path, sheet_name=sheet_name, usecols='F', skiprows=2, nrows=25)
        
        # データをリストに変換
        values_list = data.iloc[:, 0].tolist()
        
        # 物件名をファイル名から抽出
        property_name = extract_property_name(os.path.basename(file_path))
        
        # DataFrameを作成
        output_df = pd.DataFrame([values_list], columns=column_names)
        
        # 物件名の列を挿入
        output_df.insert(0, '物件名', property_name)
        
        return output_df
    except Exception as e:
        print(f"ファイル {file_path} の処理に失敗しました: {e}")
        return None

# ヒストグラムを作成して画像データを返す関数
def plot_and_save_histogram(data, title, xlabel, bins=None, color='gray', figsize=(5, 3)):
    fig, ax = plt.subplots(figsize=figsize)
    counts, bins, patches = ax.hist(data, bins=bins, edgecolor='black', alpha=0.7, color=color)
    
    # N数と平均値の計算
    mean_value = data.mean()
    sample_size = len(data)
    
    # N数と平均値の表示
    textstr = f'N={sample_size}\n平均={mean_value:.2f}'
    props = dict(boxstyle='round', facecolor='white', alpha=0.5)
    ax.text(0.95, 0.95, textstr, transform=ax.transAxes, fontsize=8,
            verticalalignment='top', horizontalalignment='right', bbox=props)

    ax.set_xlabel(xlabel, fontsize=8)
    ax.set_ylabel('頻度', fontsize=8)
    ax.grid(True)
    
    # タイトルを下に配置
    plt.figtext(0.5, -0.08, title, wrap=True, horizontalalignment='center', fontsize=10)
    
    img_data = BytesIO()
    fig.savefig(img_data, format='png', dpi=100, bbox_inches='tight')  # bbox_inches='tight' で余白を調整
    plt.close(fig)
    img_data.seek(0)
    return img_data

# ファイルパスのリストを取得
file_paths = read_file_paths(file_paths_excel)

# 全てのデータを格納するリスト
all_data = []

# 各ファイルを処理
for file_path in file_paths:
    if os.path.exists(file_path):
        print(f"処理中の物件: {os.path.basename(file_path)}")  # 処理中のファイル名を表示
        df = process_excel_file_if_exists(file_path)
        if df is not None:
            all_data.append(df)
    else:
        print(f"ファイルが存在しません: {file_path}")

# データが存在する場合の処理
if all_data:
    try:
        final_df = pd.concat(all_data, ignore_index=True)
        
        # 数値型に変換（エラーがあった場合はNaNに変換）
        final_df[column_names] = final_df[column_names].apply(pd.to_numeric, errors='coerce')
    
        final_df = final_df.merge(property_cd_df, on='物件名', how='left')
        col_order = final_df.columns.tolist()
        col_order.insert(1, col_order.pop(col_order.index('物件CD')))
        final_df = final_df[col_order]
        print(final_df[['物件名', '物件CD']].head())
        # 各カテゴリの平均値を計算し、列として追加
        for category, cols in categories.items():
            final_df[category] = final_df[cols].mean(axis=1).round(2)
        
        # 各Qw列の平均値を計算し、列として追加
        for qw, cols in qw_columns.items():
            final_df[qw] = final_df[cols].mean(axis=1).round(2)
        
        # 25項目の平均点を計算
        final_df['平均点'] = final_df[column_names].mean(axis=1).round(2)
        
        # 平均点の確認と数値型への変換
        final_df['平均点'] = pd.to_numeric(final_df['平均点'], errors='coerce')
        print("平均点の型:", final_df['平均点'].dtype)
        print("平均点の値:", final_df['平均点'])
        
        # 合計点数の計算
        final_df['合計点数'] = ((final_df['平均点'] - 3) * 25 + 50).round(2)
        
        # ランクの計算
        def calculate_rank(avg_score):
            if avg_score > 75:
                return "S"
            elif avg_score >= 65:
                return "A"
            elif avg_score >= 50:
                return "B+"
            elif avg_score >= 25:
                return "B-"
            else:
                return "C"
        
        final_df['ランク'] = final_df['合計点数'].apply(calculate_rank)
        
        # 全ての列の統計量（最大値、最小値、平均値、中央値、最頻値、分散、標準偏差）を計算
        stats_df = pd.DataFrame()
        for col in final_df.columns[1:]:
            numeric_col = pd.to_numeric(final_df[col], errors='coerce')
            stats_df[col] = {
                '最大値': numeric_col.max(),
                '最小値': numeric_col.min(),
                '平均値': numeric_col.mean(),
                '中央値': numeric_col.median(),
                '最頻値': numeric_col.mode()[0] if not numeric_col.mode().empty else None,
                '分散': numeric_col.var(),
                '標準偏差': numeric_col.std()
            }
        
        # 統計情報を転置して整形
        stats_df = stats_df.transpose().reset_index().rename(columns={'index': '項目名'})
        
        # Excelファイルを作成
        wb = Workbook()

        # データシートにデータフレームを書き込む
        ws_data = wb.active
        ws_data.title = "物件データ"
        for r in dataframe_to_rows(final_df, index=False, header=True):
            ws_data.append(r)
            
    

        # 統計情報シートを追加
        ws_stats = wb.create_sheet(title="統計情報")
        for r in dataframe_to_rows(stats_df, index=False, header=True):
            ws_stats.append(r)

        # ヒストグラムシートを追加
        ws_hist = wb.create_sheet(title="ヒストグラム")
        # 25項目のヒストグラム（1から5のスコア範囲）
        for i, col in enumerate(column_names):
            img_data = plot_and_save_histogram(final_df[col].dropna(), f' {col}', 'スコア', bins=[1, 2, 3, 4, 5, 6], color='gray')
            img = Image(img_data)
            position = f'A{1 + i * 18}'  # 位置を調整
            ws_hist.add_image(img, position)
            print(f"{col} のヒストグラムが追加されました: {position}")

        # カテゴリーに基づいたヒストグラムを追加
        for i, (category, cols) in enumerate(categories.items()):
            img_data = plot_and_save_histogram(final_df[category].dropna(), f' {category} の平均値', 'スコア', bins=[1, 2, 3, 4, 5, 6], color='gray')
            img = Image(img_data)
            position = f'I{1 + i * 18}'  # 位置を調整
            ws_hist.add_image(img, position)
            print(f"{category} のヒストグラムが追加されました: {position}")

        # Qw列に基づいたヒストグラムを追加  
        for i, (qw, cols) in enumerate(qw_columns.items()):
            img_data = plot_and_save_histogram(final_df[qw].dropna(), f' {qw} の平均値', 'スコア', bins=[1, 2, 3, 4, 5, 6], color='gray')
            img = Image(img_data)
            position = f'Q{1 + i * 18}'  # 位置を調整
            ws_hist.add_image(img, position)
            print(f"{qw} のヒストグラムが追加されました: {position}")


        # 合計点数のヒストグラム（25から85の範囲で1点刻み）
        bins = np.arange(24, 86, 1)
        img_data = plot_and_save_histogram(final_df['合計点数'].dropna(), '', '合計点数', bins=bins)
        img = Image(img_data)
        ws_hist.add_image(img, 'Y1')
        print( "合計点数のヒストグラムが追加されました: {position}")
    

        # ランクの分布（棒グラフ）
        fig, ax = plt.subplots(figsize=(5, 3))
        rank_order = ['C', 'B-', 'B+', 'A', 'S']
        final_df['ランク'].value_counts().reindex(rank_order, fill_value=0).plot(kind='bar', ax=ax, edgecolor='black', alpha=0.7, color='gray')
        ax.set_title('ランクの分布', fontsize=10)
        ax.set_xlabel('ランク', fontsize=8)
        ax.set_ylabel('頻度', fontsize=8)
        ax.grid(True)
        img_data = BytesIO()
        fig.savefig(img_data, format='png', dpi=100, bbox_inches='tight')  # bbox_inches='tight' で余白を調整
        plt.close(fig)
        img_data.seek(0)
        img = Image(img_data)
        ws_hist.add_image(img, 'Y18')  # 位置を調整
        
        
    except Exception as e:
        print(f"処理中にエラーが発生しました: {e}")

    # Excelファイルの保存
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M")
    final_output_filename = f'集計行列_{current_datetime}.xlsx'
    final_output_path = os.path.join(os.path.dirname(file_paths_excel), final_output_filename)
    wb.save(final_output_path)
    print(f"すべてのデータが次の場所に保存されました: {final_output_path}")

else:
    print("処理するデータがありませんでした。")
